from __future__ import annotations

import json
from dataclasses import fields
from pathlib import Path
from typing import Any, get_args, get_origin, get_type_hints

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from scheduler_desktop.models import (
    AppState,
    Assignment,
    Discipline,
    Room,
    ScheduleEntry,
    Stream,
    StudentGroup,
    Teacher,
)

SHEET_SPEC: list[tuple[str, str, type[Any]]] = [
    ("teachers", "teachers", Teacher),
    ("rooms", "rooms", Room),
    ("groups", "groups", StudentGroup),
    ("streams", "streams", Stream),
    ("disciplines", "disciplines", Discipline),
    ("assignments", "assignments", Assignment),
    ("schedule", "schedule", ScheduleEntry),
]


def export_state_to_excel(path: Path, state: AppState) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name, attr_name, model in SHEET_SPEC:
        ws = wb.create_sheet(title=sheet_name)
        rows = getattr(state, attr_name, [])
        _write_dataclass_rows(ws, model, rows)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def import_state_from_excel(path: Path) -> AppState:
    wb = load_workbook(path, data_only=True)
    data: dict[str, Any] = {}

    for sheet_name, attr_name, model in SHEET_SPEC:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data[attr_name] = _read_dataclass_rows(ws, model)
        else:
            data[attr_name] = []

    return AppState(**data)


def _write_dataclass_rows(ws: Worksheet, model: type[Any], rows: list[Any]) -> None:
    cols = [f.name for f in fields(model)]
    ws.append(cols)
    for row in rows:
        ws.append([_serialize_cell(getattr(row, col)) for col in cols])


def _read_dataclass_rows(ws: Worksheet, model: type[Any]) -> list[Any]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    if not any(headers):
        return []

    allowed = {f.name for f in fields(model)}
    hints = get_type_hints(model)
    result: list[Any] = []

    for row in rows[1:]:
        if not any(cell not in (None, "") for cell in row):
            continue
        record: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header or header not in allowed:
                continue
            raw = row[idx] if idx < len(row) else None
            if raw in (None, ""):
                continue
            record[header] = _deserialize_cell(raw, hints.get(header, str))
        result.append(model(**record))

    return result


def _serialize_cell(value: Any) -> Any:
    if isinstance(value, list):
        return json.dumps(value, ensure_ascii=False)
    return value


def _deserialize_cell(value: Any, target_type: Any) -> Any:
    origin = get_origin(target_type)
    args = get_args(target_type)

    if origin is list:
        return _parse_list(value)

    if origin is None and target_type is list:
        return _parse_list(value)

    if origin is None and target_type is bool:
        return _parse_bool(value)

    if origin is None and target_type is int:
        return int(value)

    if origin is not None and type(None) in args:
        non_none = [a for a in args if a is not type(None)]
        if len(non_none) == 1:
            inner = non_none[0]
            if inner is int:
                return int(value)
            if inner is bool:
                return _parse_bool(value)
            if get_origin(inner) is list:
                return _parse_list(value)
        return value

    return value


def _parse_list(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        if text.startswith("["):
            parsed = json.loads(text)
            if isinstance(parsed, list):
                return parsed
        return [item.strip() for item in text.split(",") if item.strip()]
    return [value]


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "да", "y"}
